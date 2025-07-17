"""
Multi-Format Export Module for Industrial Compliance
Supports JSON, CSV, XML, SARIF, STIX, and other industry-standard formats
"""

import json
import csv
import xml.etree.ElementTree as ET
import xml.dom.minidom
from datetime import datetime
from pathlib import Path
import uuid
import os

class MultiFormatExporter:
    """Export security findings to multiple industrial formats"""
    
    def __init__(self):
        self.supported_formats = [
            'json', 'csv', 'xml', 'sarif', 'stix', 'mitre', 'nist',
            'excel', 'markdown', 'txt', 'yaml', 'junit'
        ]
    
    def export(self, findings, output_path, format_type='json', metadata=None):
        """
        Export findings to specified format
        
        Args:
            findings (list): List of security findings
            output_path (str): Output file path
            format_type (str): Export format type
            metadata (dict): Additional metadata
        
        Returns:
            str: Path to exported file
        """
        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Set default metadata
        if not metadata:
            metadata = self.get_default_metadata()
        
        # Export based on format type
        if format_type == 'json':
            return self.export_json(findings, output_path, metadata)
        elif format_type == 'csv':
            return self.export_csv(findings, output_path, metadata)
        elif format_type == 'xml':
            return self.export_xml(findings, output_path, metadata)
        elif format_type == 'sarif':
            return self.export_sarif(findings, output_path, metadata)
        elif format_type == 'stix':
            return self.export_stix(findings, output_path, metadata)
        elif format_type == 'mitre':
            return self.export_mitre(findings, output_path, metadata)
        elif format_type == 'nist':
            return self.export_nist(findings, output_path, metadata)
        elif format_type == 'excel':
            return self.export_excel(findings, output_path, metadata)
        elif format_type == 'markdown':
            return self.export_markdown(findings, output_path, metadata)
        elif format_type == 'txt':
            return self.export_txt(findings, output_path, metadata)
        elif format_type == 'yaml':
            return self.export_yaml(findings, output_path, metadata)
        elif format_type == 'junit':
            return self.export_junit(findings, output_path, metadata)
        else:
            raise ValueError(f"Unsupported format: {format_type}")
    
    def export_json(self, findings, output_path, metadata):
        """Export to JSON format"""
        try:
            report_data = {
                'metadata': metadata,
                'summary': self.generate_summary(findings),
                'findings': findings,
                'statistics': self.generate_statistics(findings),
                'generated_at': datetime.now().isoformat(),
                'schema_version': '1.0'
            }
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=2, ensure_ascii=False)
            
            print(f"[OK] JSON export completed: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"[ERROR] JSON export failed: {str(e)}")
            raise
    
    def export_csv(self, findings, output_path, metadata):
        """Export to CSV format"""
        try:
            fieldnames = [
                'id', 'title', 'description', 'severity', 'cvss_score',
                'category', 'subcategory', 'host', 'port', 'service',
                'impact', 'evidence', 'remediation', 'source',
                'discovery_date', 'last_seen', 'confidence'
            ]
            
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                
                for finding in findings:
                    # Flatten finding data for CSV
                    csv_row = {
                        'id': finding.get('id', str(uuid.uuid4())),
                        'title': finding.get('title', ''),
                        'description': finding.get('description', ''),
                        'severity': finding.get('severity', ''),
                        'cvss_score': finding.get('cvss_score', ''),
                        'category': finding.get('category', ''),
                        'subcategory': finding.get('subcategory', ''),
                        'host': finding.get('host', ''),
                        'port': finding.get('port', ''),
                        'service': finding.get('service', ''),
                        'impact': finding.get('impact', ''),
                        'evidence': finding.get('evidence', ''),
                        'remediation': finding.get('remediation', ''),
                        'source': finding.get('source', ''),
                        'discovery_date': finding.get('discovery_date', ''),
                        'last_seen': finding.get('last_seen', ''),
                        'confidence': finding.get('confidence', '')
                    }
                    writer.writerow(csv_row)
            
            print(f"[OK] CSV export completed: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"[ERROR] CSV export failed: {str(e)}")
            raise
    
    def export_xml(self, findings, output_path, metadata):
        """Export to XML format"""
        try:
            root = ET.Element('security_report')
            
            # Add metadata
            metadata_elem = ET.SubElement(root, 'metadata')
            for key, value in metadata.items():
                elem = ET.SubElement(metadata_elem, key)
                elem.text = str(value)
            
            # Add summary
            summary_elem = ET.SubElement(root, 'summary')
            summary_data = self.generate_summary(findings)
            for key, value in summary_data.items():
                elem = ET.SubElement(summary_elem, key)
                elem.text = str(value)
            
            # Add findings
            findings_elem = ET.SubElement(root, 'findings')
            for finding in findings:
                finding_elem = ET.SubElement(findings_elem, 'finding')
                finding_elem.set('id', finding.get('id', str(uuid.uuid4())))
                
                for key, value in finding.items():
                    if isinstance(value, (list, dict)):
                        # Handle complex data types
                        if isinstance(value, list):
                            list_elem = ET.SubElement(finding_elem, key)
                            for item in value:
                                item_elem = ET.SubElement(list_elem, 'item')
                                item_elem.text = str(item)
                        else:
                            dict_elem = ET.SubElement(finding_elem, key)
                            for k, v in value.items():
                                sub_elem = ET.SubElement(dict_elem, k)
                                sub_elem.text = str(v)
                    else:
                        elem = ET.SubElement(finding_elem, key)
                        elem.text = str(value) if value is not None else ''
            
            # Format and save XML
            xml_str = ET.tostring(root, encoding='utf-8')
            dom = xml.dom.minidom.parseString(xml_str)
            pretty_xml = dom.toprettyxml(indent='  ')
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(pretty_xml)
            
            print(f"[OK] XML export completed: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"[ERROR] XML export failed: {str(e)}")
            raise
    
    def export_sarif(self, findings, output_path, metadata):
        """Export to SARIF (Static Analysis Results Interchange Format)"""
        try:
            # SARIF 2.1.0 format
            sarif_report = {
                'version': '2.1.0',
                '$schema': 'https://raw.githubusercontent.com/oasis-tcs/sarif-spec/master/Schemata/sarif-schema-2.1.0.json',
                'runs': [
                    {
                        'tool': {
                            'driver': {
                                'name': 'CyberSec-AI AutoReport',
                                'version': '2.0.0',
                                'organization': metadata.get('organization', 'Security Team'),
                                'rules': self.generate_sarif_rules(findings)
                            }
                        },
                        'results': self.generate_sarif_results(findings),
                        'invocations': [
                            {
                                'executionSuccessful': True,
                                'endTimeUtc': datetime.now().isoformat() + 'Z'
                            }
                        ]
                    }
                ]
            }
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(sarif_report, f, indent=2, ensure_ascii=False)
            
            print(f"[OK] SARIF export completed: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"[ERROR] SARIF export failed: {str(e)}")
            raise
    
    def export_stix(self, findings, output_path, metadata):
        """Export to STIX (Structured Threat Information Expression)"""
        try:
            # STIX 2.1 format
            stix_bundle = {
                'type': 'bundle',
                'id': f"bundle--{uuid.uuid4()}",
                'spec_version': '2.1',
                'objects': []
            }
            
            # Add identity object
            identity = {
                'type': 'identity',
                'id': f"identity--{uuid.uuid4()}",
                'created': datetime.now().isoformat() + 'Z',
                'modified': datetime.now().isoformat() + 'Z',
                'name': metadata.get('organization', 'Security Team'),
                'identity_class': 'organization'
            }
            stix_bundle['objects'].append(identity)
            
            # Add vulnerability objects
            for finding in findings:
                vulnerability = {
                    'type': 'vulnerability',
                    'id': f"vulnerability--{uuid.uuid4()}",
                    'created': datetime.now().isoformat() + 'Z',
                    'modified': datetime.now().isoformat() + 'Z',
                    'name': finding.get('title', 'Unknown Vulnerability'),
                    'description': finding.get('description', ''),
                    'labels': [finding.get('category', 'vulnerability').lower()],
                    'x_severity': finding.get('severity', 'medium').lower(),
                    'x_cvss_score': finding.get('cvss_score', 0),
                    'x_affected_assets': [finding.get('host', 'unknown')]
                }
                stix_bundle['objects'].append(vulnerability)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(stix_bundle, f, indent=2, ensure_ascii=False)
            
            print(f"[OK] STIX export completed: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"[ERROR] STIX export failed: {str(e)}")
            raise
    
    def export_mitre(self, findings, output_path, metadata):
        """Export to MITRE ATT&CK framework format"""
        try:
            mitre_report = {
                'metadata': metadata,
                'attack_patterns': [],
                'techniques': [],
                'tactics': [],
                'findings_mapping': []
            }
            
            # Map findings to MITRE ATT&CK techniques
            for finding in findings:
                # Simplified mapping based on vulnerability types
                techniques = self.map_to_mitre_techniques(finding)
                
                finding_mapping = {
                    'finding_id': finding.get('id', str(uuid.uuid4())),
                    'finding_title': finding.get('title', ''),
                    'mapped_techniques': techniques,
                    'confidence': finding.get('confidence', 'medium')
                }
                mitre_report['findings_mapping'].append(finding_mapping)
                
                # Add unique techniques
                for technique in techniques:
                    if technique not in mitre_report['techniques']:
                        mitre_report['techniques'].append(technique)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(mitre_report, f, indent=2, ensure_ascii=False)
            
            print(f"[OK] MITRE ATT&CK export completed: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"[ERROR] MITRE export failed: {str(e)}")
            raise
    
    def export_nist(self, findings, output_path, metadata):
        """Export to NIST Cybersecurity Framework format"""
        try:
            nist_report = {
                'metadata': metadata,
                'framework_version': '1.1',
                'functions': {
                    'identify': {'subcategories': []},
                    'protect': {'subcategories': []},
                    'detect': {'subcategories': []},
                    'respond': {'subcategories': []},
                    'recover': {'subcategories': []}
                },
                'findings_mapping': []
            }
            
            # Map findings to NIST subcategories
            for finding in findings:
                nist_mapping = self.map_to_nist_subcategories(finding)
                
                finding_mapping = {
                    'finding_id': finding.get('id', str(uuid.uuid4())),
                    'finding_title': finding.get('title', ''),
                    'severity': finding.get('severity', ''),
                    'nist_subcategories': nist_mapping,
                    'recommendations': finding.get('remediation', '')
                }
                nist_report['findings_mapping'].append(finding_mapping)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(nist_report, f, indent=2, ensure_ascii=False)
            
            print(f"[OK] NIST Framework export completed: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"[ERROR] NIST export failed: {str(e)}")
            raise
    
    def export_excel(self, findings, output_path, metadata):
        """Export to Excel format"""
        try:
            # Try to use openpyxl if available
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                wb = Workbook()
                
                # Summary sheet
                ws_summary = wb.active
                ws_summary.title = 'Executive Summary'
                
                # Add summary data
                summary_data = self.generate_summary(findings)
                row = 1
                for key, value in summary_data.items():
                    ws_summary.cell(row=row, column=1, value=key.replace('_', ' ').title())
                    ws_summary.cell(row=row, column=2, value=str(value))
                    row += 1
                
                # Findings sheet
                ws_findings = wb.create_sheet('Detailed Findings')
                
                # Headers
                headers = [
                    'ID', 'Title', 'Description', 'Severity', 'CVSS Score',
                    'Category', 'Host', 'Port', 'Service', 'Impact',
                    'Evidence', 'Remediation', 'Source', 'Discovery Date'
                ]
                
                for col, header in enumerate(headers, 1):
                    cell = ws_findings.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                
                # Data rows
                for row, finding in enumerate(findings, 2):
                    ws_findings.cell(row=row, column=1, value=finding.get('id', ''))
                    ws_findings.cell(row=row, column=2, value=finding.get('title', ''))
                    ws_findings.cell(row=row, column=3, value=finding.get('description', ''))
                    ws_findings.cell(row=row, column=4, value=finding.get('severity', ''))
                    ws_findings.cell(row=row, column=5, value=finding.get('cvss_score', ''))
                    ws_findings.cell(row=row, column=6, value=finding.get('category', ''))
                    ws_findings.cell(row=row, column=7, value=finding.get('host', ''))
                    ws_findings.cell(row=row, column=8, value=finding.get('port', ''))
                    ws_findings.cell(row=row, column=9, value=finding.get('service', ''))
                    ws_findings.cell(row=row, column=10, value=finding.get('impact', ''))
                    ws_findings.cell(row=row, column=11, value=finding.get('evidence', ''))
                    ws_findings.cell(row=row, column=12, value=finding.get('remediation', ''))
                    ws_findings.cell(row=row, column=13, value=finding.get('source', ''))
                    ws_findings.cell(row=row, column=14, value=finding.get('discovery_date', ''))
                
                wb.save(output_path)
                
            except ImportError:
                # Fallback to CSV if openpyxl not available
                return self.export_csv(findings, output_path.replace('.xlsx', '.csv'), metadata)
            
            print(f"[OK] Excel export completed: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"[ERROR] Excel export failed: {str(e)}")
            raise
    
    def export_markdown(self, findings, output_path, metadata):
        """Export to Markdown format"""
        try:
            md_content = f"""# {metadata.get('title', 'Cybersecurity Assessment Report')}

## Executive Summary

**Organization:** {metadata.get('organization', 'N/A')}  
**Assessment Date:** {metadata.get('assessment_date', 'N/A')}  
**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  
**Total Findings:** {len(findings)}

## Summary Statistics

"""
            
            # Add statistics
            stats = self.generate_statistics(findings)
            md_content += f"| Severity | Count |\n|----------|-------|\n"
            for severity, count in stats.get('by_severity', {}).items():
                md_content += f"| {severity.title()} | {count} |\n"
            
            md_content += "\n## Detailed Findings\n\n"
            
            # Add findings
            for i, finding in enumerate(findings, 1):
                md_content += f"### {i}. {finding.get('title', 'Unknown Vulnerability')}\n\n"
                md_content += f"**Severity:** {finding.get('severity', 'Unknown')}  \n"
                md_content += f"**CVSS Score:** {finding.get('cvss_score', 'N/A')}  \n"
                md_content += f"**Category:** {finding.get('category', 'Unknown')}  \n"
                md_content += f"**Host:** {finding.get('host', 'N/A')}  \n"
                md_content += f"**Port:** {finding.get('port', 'N/A')}  \n\n"
                
                md_content += f"**Description:**  \n{finding.get('description', 'No description available')}\n\n"
                
                if finding.get('impact'):
                    md_content += f"**Impact:**  \n{finding.get('impact')}\n\n"
                
                if finding.get('evidence'):
                    md_content += f"**Evidence:**  \n```\n{finding.get('evidence')}\n```\n\n"
                
                if finding.get('remediation'):
                    md_content += f"**Remediation:**  \n{finding.get('remediation')}\n\n"
                
                md_content += "---\n\n"
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(md_content)
            
            print(f"[OK] Markdown export completed: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"[ERROR] Markdown export failed: {str(e)}")
            raise
    
    def export_txt(self, findings, output_path, metadata):
        """Export to plain text format"""
        try:
            txt_content = f"""{metadata.get('title', 'CYBERSECURITY ASSESSMENT REPORT')}
{'=' * 60}

Organization: {metadata.get('organization', 'N/A')}
Assessment Date: {metadata.get('assessment_date', 'N/A')}
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Total Findings: {len(findings)}

EXECUTIVE SUMMARY
{'-' * 17}

This report contains {len(findings)} security findings identified during the assessment.

"""
            
            # Add statistics
            stats = self.generate_statistics(findings)
            txt_content += "SEVERITY DISTRIBUTION\n"
            txt_content += "-" * 20 + "\n"
            for severity, count in stats.get('by_severity', {}).items():
                txt_content += f"{severity.title():<10}: {count}\n"
            
            txt_content += "\nDETAILED FINDINGS\n"
            txt_content += "=" * 17 + "\n\n"
            
            # Add findings
            for i, finding in enumerate(findings, 1):
                txt_content += f"{i}. {finding.get('title', 'Unknown Vulnerability')}\n"
                txt_content += "-" * (len(str(i)) + 2 + len(finding.get('title', 'Unknown Vulnerability'))) + "\n"
                txt_content += f"Severity: {finding.get('severity', 'Unknown')}\n"
                txt_content += f"CVSS Score: {finding.get('cvss_score', 'N/A')}\n"
                txt_content += f"Category: {finding.get('category', 'Unknown')}\n"
                txt_content += f"Host: {finding.get('host', 'N/A')}\n"
                txt_content += f"Port: {finding.get('port', 'N/A')}\n\n"
                
                txt_content += f"Description:\n{finding.get('description', 'No description available')}\n\n"
                
                if finding.get('impact'):
                    txt_content += f"Impact:\n{finding.get('impact')}\n\n"
                
                if finding.get('evidence'):
                    txt_content += f"Evidence:\n{finding.get('evidence')}\n\n"
                
                if finding.get('remediation'):
                    txt_content += f"Remediation:\n{finding.get('remediation')}\n\n"
                
                txt_content += "=" * 60 + "\n\n"
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(txt_content)
            
            print(f"[OK] Text export completed: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"[ERROR] Text export failed: {str(e)}")
            raise
    
    def export_yaml(self, findings, output_path, metadata):
        """Export to YAML format"""
        try:
            import yaml
            
            report_data = {
                'metadata': metadata,
                'summary': self.generate_summary(findings),
                'findings': findings,
                'statistics': self.generate_statistics(findings),
                'generated_at': datetime.now().isoformat()
            }
            
            with open(output_path, 'w', encoding='utf-8') as f:
                yaml.dump(report_data, f, default_flow_style=False, allow_unicode=True)
            
            print(f"[OK] YAML export completed: {output_path}")
            return output_path
            
        except ImportError:
            print("[WARNING] PyYAML not available, falling back to JSON")
            return self.export_json(findings, output_path.replace('.yaml', '.json'), metadata)
        except Exception as e:
            print(f"[ERROR] YAML export failed: {str(e)}")
            raise
    
    def export_junit(self, findings, output_path, metadata):
        """Export to JUnit XML format for CI/CD integration"""
        try:
            root = ET.Element('testsuites')
            root.set('name', 'Security Assessment')
            root.set('tests', str(len(findings)))
            root.set('failures', str(len([f for f in findings if f.get('severity', '').lower() in ['critical', 'high']])))
            root.set('time', '0')
            
            testsuite = ET.SubElement(root, 'testsuite')
            testsuite.set('name', 'Vulnerability Scan')
            testsuite.set('tests', str(len(findings)))
            testsuite.set('failures', str(len([f for f in findings if f.get('severity', '').lower() in ['critical', 'high']])))
            testsuite.set('time', '0')
            
            for finding in findings:
                testcase = ET.SubElement(testsuite, 'testcase')
                testcase.set('name', finding.get('title', 'Unknown Vulnerability'))
                testcase.set('classname', finding.get('category', 'Security'))
                testcase.set('time', '0')
                
                # Mark high/critical as failures
                if finding.get('severity', '').lower() in ['critical', 'high']:
                    failure = ET.SubElement(testcase, 'failure')
                    failure.set('message', finding.get('description', ''))
                    failure.set('type', finding.get('severity', ''))
                    failure.text = finding.get('remediation', 'No remediation provided')
            
            # Format and save XML
            xml_str = ET.tostring(root, encoding='utf-8')
            dom = xml.dom.minidom.parseString(xml_str)
            pretty_xml = dom.toprettyxml(indent='  ')
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(pretty_xml)
            
            print(f"[OK] JUnit export completed: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"[ERROR] JUnit export failed: {str(e)}")
            raise
    
    def generate_summary(self, findings):
        """Generate executive summary statistics"""
        total = len(findings)
        by_severity = {}
        
        for finding in findings:
            severity = finding.get('severity', 'unknown').lower()
            by_severity[severity] = by_severity.get(severity, 0) + 1
        
        return {
            'total_findings': total,
            'critical_findings': by_severity.get('critical', 0),
            'high_findings': by_severity.get('high', 0),
            'medium_findings': by_severity.get('medium', 0),
            'low_findings': by_severity.get('low', 0),
            'info_findings': by_severity.get('info', 0)
        }
    
    def generate_statistics(self, findings):
        """Generate detailed statistics"""
        stats = {
            'total': len(findings),
            'by_severity': {},
            'by_category': {},
            'by_host': {},
            'by_source': {}
        }
        
        for finding in findings:
            # By severity
            severity = finding.get('severity', 'unknown').lower()
            stats['by_severity'][severity] = stats['by_severity'].get(severity, 0) + 1
            
            # By category
            category = finding.get('category', 'unknown')
            stats['by_category'][category] = stats['by_category'].get(category, 0) + 1
            
            # By host
            host = finding.get('host', 'unknown')
            stats['by_host'][host] = stats['by_host'].get(host, 0) + 1
            
            # By source
            source = finding.get('source', 'unknown')
            stats['by_source'][source] = stats['by_source'].get(source, 0) + 1
        
        return stats
    
    def generate_sarif_rules(self, findings):
        """Generate SARIF rules from findings"""
        rules = []
        rule_ids = set()
        
        for finding in findings:
            rule_id = f"rule-{finding.get('category', 'general').lower().replace(' ', '-')}"
            
            if rule_id not in rule_ids:
                rule = {
                    'id': rule_id,
                    'name': finding.get('category', 'General Security Rule'),
                    'shortDescription': {
                        'text': finding.get('title', 'Security vulnerability detected')
                    },
                    'fullDescription': {
                        'text': finding.get('description', 'No description available')
                    },
                    'defaultConfiguration': {
                        'level': self.severity_to_sarif_level(finding.get('severity', 'medium'))
                    },
                    'helpUri': 'https://example.com/help'
                }
                rules.append(rule)
                rule_ids.add(rule_id)
        
        return rules
    
    def generate_sarif_results(self, findings):
        """Generate SARIF results from findings"""
        results = []
        
        for finding in findings:
            result = {
                'ruleId': f"rule-{finding.get('category', 'general').lower().replace(' ', '-')}",
                'level': self.severity_to_sarif_level(finding.get('severity', 'medium')),
                'message': {
                    'text': finding.get('title', 'Security vulnerability detected')
                },
                'locations': [
                    {
                        'physicalLocation': {
                            'artifactLocation': {
                                'uri': finding.get('host', 'unknown')
                            }
                        }
                    }
                ]
            }
            results.append(result)
        
        return results
    
    def severity_to_sarif_level(self, severity):
        """Convert severity to SARIF level"""
        mapping = {
            'critical': 'error',
            'high': 'error',
            'medium': 'warning',
            'low': 'note',
            'info': 'note'
        }
        return mapping.get(severity.lower(), 'warning')
    
    def map_to_mitre_techniques(self, finding):
        """Map finding to MITRE ATT&CK techniques"""
        # Simplified mapping based on common vulnerability types
        category = finding.get('category', '').lower()
        title = finding.get('title', '').lower()
        
        techniques = []
        
        if 'web' in category or 'application' in category:
            techniques.append('T1190')  # Exploit Public-Facing Application
        if 'network' in category:
            techniques.append('T1046')  # Network Service Scanning
        if 'system' in category:
            techniques.append('T1068')  # Exploitation for Privilege Escalation
        if 'sql' in title or 'injection' in title:
            techniques.append('T1190')  # Exploit Public-Facing Application
        if 'xss' in title or 'cross-site' in title:
            techniques.append('T1190')  # Exploit Public-Facing Application
        
        return techniques if techniques else ['T1190']  # Default to exploitation
    
    def map_to_nist_subcategories(self, finding):
        """Map finding to NIST Cybersecurity Framework subcategories"""
        # Simplified mapping
        category = finding.get('category', '').lower()
        severity = finding.get('severity', '').lower()
        
        subcategories = []
        
        if 'web' in category or 'application' in category:
            subcategories.extend(['PR.AC-4', 'PR.DS-5', 'DE.CM-1'])
        if 'network' in category:
            subcategories.extend(['PR.AC-5', 'PR.PT-3', 'DE.CM-1'])
        if 'system' in category:
            subcategories.extend(['PR.AC-1', 'PR.PT-1', 'DE.CM-7'])
        
        # Add response subcategories for high/critical
        if severity in ['critical', 'high']:
            subcategories.extend(['RS.RP-1', 'RS.MI-1'])
        
        return subcategories if subcategories else ['ID.AM-1']  # Default to asset management
    
    def get_default_metadata(self):
        """Get default metadata for exports"""
        return {
            'title': 'Cybersecurity Assessment Report',
            'organization': 'Organization',
            'assessment_date': datetime.now().strftime('%Y-%m-%d'),
            'generated_by': 'CyberSec-AI AutoReport',
            'version': '2.0',
            'assessment_type': 'Automated Security Assessment',
            'scope': 'Network and Application Security',
            'methodology': 'Vulnerability Scanning and Analysis'
        }
    
    def get_supported_formats(self):
        """Get list of supported export formats"""
        return self.supported_formats
    
    def validate_format(self, format_type):
        """Validate if format is supported"""
        return format_type in self.supported_formats

# Convenience functions
def export_to_multiple_formats(findings, base_path, formats=['json', 'csv', 'xml'], metadata=None):
    """Export findings to multiple formats"""
    exporter = MultiFormatExporter()
    exported_files = []
    
    for format_type in formats:
        try:
            file_extension = 'xlsx' if format_type == 'excel' else format_type
            output_path = f"{base_path}.{file_extension}"
            exported_file = exporter.export(findings, output_path, format_type, metadata)
            exported_files.append(exported_file)
        except Exception as e:
            print(f"[ERROR] Failed to export to {format_type}: {str(e)}")
    
    return exported_files

if __name__ == "__main__":
    # Test the multi-format exporter
    test_findings = [
        {
            'id': str(uuid.uuid4()),
            'title': 'SQL Injection Vulnerability',
            'description': 'SQL injection vulnerability found in login form',
            'severity': 'High',
            'category': 'Web Application',
            'host': 'example.com',
            'port': '80',
            'service': 'HTTP',
            'evidence': 'Error-based SQL injection confirmed',
            'impact': 'Unauthorized database access',
            'remediation': 'Use parameterized queries'
        }
    ]
    
    exporter = MultiFormatExporter()
    
    print("Supported formats:")
    for fmt in exporter.get_supported_formats():
        print(f"  - {fmt}")
    
    # Test JSON export
    json_path = exporter.export(test_findings, "test_report.json", "json")
    print(f"Test JSON export: {json_path}")
